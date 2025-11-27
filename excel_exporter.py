"""
Excel Exporter Module
Exports extracted resume data to Excel format.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import List, Dict, Any
from datetime import datetime


def export_to_excel(data: List[Dict[str, Any]]) -> bytes:
    """
    Export extraction results to Excel file.
    
    Args:
        data: List of extraction results with keys:
            - name: Candidate name
            - email: Email address
            - phone: Phone number
            - filename: Source file name
            - error: Error message if any
            
    Returns:
        Excel file as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Headers
    headers = ["Names", "Email", "Phone Numbers", "FileName"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_idx, item in enumerate(data, 2):
        # Names
        name_cell = ws.cell(row=row_idx, column=1, value=item.get('name') or "Not found")
        name_cell.alignment = cell_alignment
        name_cell.border = thin_border
        
        # Email
        email_cell = ws.cell(row=row_idx, column=2, value=item.get('email') or "Not found")
        email_cell.alignment = cell_alignment
        email_cell.border = thin_border
        
        # Phone Numbers
        phone_cell = ws.cell(row=row_idx, column=3, value=item.get('phone') or "Not found")
        phone_cell.alignment = cell_alignment
        phone_cell.border = thin_border
        
        # FileName (include error if any)
        filename = item.get('filename', 'Unknown')
        error = item.get('error')
        if error:
            filename = f"{filename} (Error: {error})"
        
        file_cell = ws.cell(row=row_idx, column=4, value=filename)
        file_cell.alignment = cell_alignment
        file_cell.border = thin_border
        
        # Highlight rows with errors
        if error:
            error_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = error_fill
    
    # Auto-adjust column widths
    column_widths = [30, 35, 20, 45]  # Approximate widths
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def get_export_filename() -> str:
    """Generate export filename with current date."""
    return f"resume_extraction.xlsx"
